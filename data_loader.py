# DATA LOADER — reads Excel files from OneDrive via Microsoft Graph API or local disk
import os, datetime, tempfile
from io import BytesIO
from openpyxl import load_workbook
from collections import defaultdict

try:
    import requests
except ImportError:
    requests = None

BILLING_MONTHS = [
    'Jan-26','Feb-26','Mar-26','Apr-26','May-26','Jun-26',
    'Jul-26','Aug-26','Sept-26','Oct-26','Nov-26','Dec-26'
]

def _get_year(val):
    if val is None: return ""
    if isinstance(val, datetime.datetime): return str(val.year)
    if isinstance(val, (int, float)): return str(int(val))
    return ""

def _fmt_bm(val):
    if val is None: return ""
    if isinstance(val, datetime.datetime): return val.strftime("%b-%y")
    if isinstance(val, (int, float)): return str(int(val))
    return str(val).strip()

def _clean_status(val):
    if val is None: return ""
    if isinstance(val, (int, float)): return "{}%".format(int(round(val * 100)))
    if isinstance(val, datetime.datetime): return val.strftime("%d-%b-%Y")
    return str(val).strip()

def _get_token(tenant_id, client_id, client_secret):
    url = "https://login.microsoftonline.com/{}/oauth2/v2.0/token".format(tenant_id)
    resp = requests.post(url, data={
        "grant_type":    "client_credentials",
        "client_id":     client_id,
        "client_secret": client_secret,
        "scope":         "https://graph.microsoft.com/.default",
    }, timeout=15)
    resp.raise_for_status()
    return resp.json()["access_token"]

def _download_file(token, user_email, folder_path, filename):
    path = "{}/{}".format(folder_path.rstrip("/"), filename)
    url  = "https://graph.microsoft.com/v1.0/users/{}/drive/root:/{}:/content".format(
        user_email, path)
    resp = requests.get(url, headers={"Authorization": "Bearer " + token}, timeout=30)
    resp.raise_for_status()
    return resp.content  # bytes

def load_dashboard_data(data_folder, tracker_files, onedrive_cfg=None):
    errors      = []
    targets     = {}
    bb_rows     = []
    pipeline    = defaultdict(lambda: defaultdict(float))
    pipe_detail = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    billing_raw = defaultdict(lambda: defaultdict(float))   # {label: {month: value}}
    am_order    = list(tracker_files.keys())

    # ── Decide: OneDrive mode or local file mode ──────────────────────────────
    use_onedrive = (
        onedrive_cfg and requests and
        onedrive_cfg.get("tenant_id") and
        onedrive_cfg.get("client_id") and
        onedrive_cfg.get("client_secret") and
        onedrive_cfg.get("user_email") and
        onedrive_cfg.get("folder_path")
    )

    token = None
    if use_onedrive:
        try:
            token = _get_token(
                onedrive_cfg["tenant_id"],
                onedrive_cfg["client_id"],
                onedrive_cfg["client_secret"],
            )
        except Exception as e:
            errors.append("OneDrive auth failed: {}".format(e))
            use_onedrive = False

    for label, fname in tracker_files.items():
        wb = None
        try:
            if use_onedrive and token:
                file_bytes = _download_file(
                    token, onedrive_cfg["user_email"],
                    onedrive_cfg["folder_path"], fname,
                )
                wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
            else:
                fpath = os.path.join(data_folder, fname)
                if not os.path.exists(fpath):
                    errors.append("File not found: " + fpath)
                    targets[label] = 0
                    continue
                wb = load_workbook(fpath, read_only=True, data_only=True)

            targets[label] = 0

            # ── Summary: annual target ────────────────────────────────────────
            if "Summary" in wb.sheetnames:
                for row in wb["Summary"].iter_rows(values_only=True):
                    if len(row) > 2 and row[1] == "Top Line Target 2026":
                        v = row[2]
                        targets[label] = float(v) if isinstance(v, (int, float)) else 0
                        break

            # ── Booking & Billing Forecast ────────────────────────────────────
            if "Booking & Billing Forecast" in wb.sheetnames:
                for i, row in enumerate(wb["Booking & Billing Forecast"].iter_rows(values_only=True)):
                    if i < 2 or len(row) < 15: continue
                    acct, cust, order, bm = row[1], row[3], row[6], row[14]
                    if not acct or not order: continue
                    if isinstance(acct, str) and not acct.strip(): continue
                    if isinstance(order, str) and (not order.strip() or order.startswith("=")): continue
                    try: v = float(order)
                    except: continue
                    cust_str = str(cust).strip() if cust else "-"
                    bb_rows.append((label, cust_str, v, _fmt_bm(bm), _get_year(bm)))

                    # Monthly invoice values — col 16, 19, 22 … (every 3 columns)
                    for m_idx, mname in enumerate(BILLING_MONTHS):
                        col = 16 + m_idx * 3
                        if col < len(row) and row[col]:
                            try:
                                mv = float(row[col])
                                if mv > 0:
                                    billing_raw[label][mname] += mv
                            except: pass

            # ── Opportunity Tracker ───────────────────────────────────────────
            if "Opportunity Tracker" in wb.sheetnames:
                for i, row in enumerate(wb["Opportunity Tracker"].iter_rows(values_only=True)):
                    if i < 1 or len(row) < 12: continue
                    acct, cust, prop, status = row[1], row[3], row[7], row[11]
                    if not acct or not prop: continue
                    if isinstance(acct, str) and not acct.strip(): continue
                    if isinstance(prop, str) and (not prop.strip() or prop.startswith("=")): continue
                    try: pv = float(prop)
                    except: continue
                    sc = _clean_status(status)
                    if sc in ("40%", "50%", "80%"):
                        pipeline[label][sc] += pv
                        cust_name = str(cust).strip() if cust else "-"
                        pipe_detail[label][sc][cust_name] += pv

            wb.close()

        except Exception as e:
            errors.append("Error reading {}: {}".format(fname, e))
            targets.setdefault(label, 0)

    # ── Aggregate bookings ────────────────────────────────────────────────────
    actual_2026 = defaultdict(float)
    cust_totals = defaultdict(lambda: defaultdict(float))
    for label, cust, v, bm, yr in bb_rows:
        if yr == "2026":
            actual_2026[label] += v
            cust_totals[label][cust] += v

    accounts_2026 = {}
    for label in am_order:
        consolidated = sorted(cust_totals[label].items(), key=lambda x: -x[1])
        accounts_2026[label] = [{"customer": c, "amount": round(t, 2)} for c, t in consolidated]

    pipeline_detail = {}
    for label in am_order:
        pipeline_detail[label] = {}
        for sc in ("40%", "50%", "80%"):
            items = sorted(pipe_detail[label][sc].items(), key=lambda x: -x[1])
            pipeline_detail[label][sc] = [{"customer": c, "amount": round(v, 2)} for c, v in items]

    # ── Billing monthly ───────────────────────────────────────────────────────
    billing_monthly = {}
    total_billed_2026 = {}
    for label in am_order:
        billing_monthly[label] = {
            mname: round(billing_raw[label].get(mname, 0), 2)
            for mname in BILLING_MONTHS
        }
        total_billed_2026[label] = round(sum(billing_monthly[label].values()), 2)

    total_target      = sum(targets.values())
    total_actual_2026 = sum(actual_2026.values())
    grand_billed_2026 = sum(total_billed_2026.values())

    # last_updated
    if use_onedrive:
        lu_str = datetime.datetime.utcnow().strftime("%d %b %Y, %H:%M UTC")
    else:
        last_updated = None
        for fname in tracker_files.values():
            fpath = os.path.join(data_folder, fname)
            if os.path.exists(fpath):
                mtime = datetime.datetime.fromtimestamp(os.path.getmtime(fpath))
                if last_updated is None or mtime > last_updated:
                    last_updated = mtime
        lu_str = last_updated.strftime("%d %b %Y, %H:%M") if last_updated else "Unknown"

    return {
        "total_target":       total_target,
        "total_actual_2026":  total_actual_2026,
        "grand_billed_2026":  grand_billed_2026,
        "targets":            targets,
        "actual_2026":        dict(actual_2026),
        "accounts_2026":      accounts_2026,
        "pipeline":           {am: dict(pipeline[am]) for am in am_order},
        "pipeline_detail":    pipeline_detail,
        "billing_monthly":    billing_monthly,
        "total_billed_2026":  total_billed_2026,
        "billing_months":     BILLING_MONTHS,
        "am_labels":          am_order,
        "last_updated":       lu_str,
        "errors":             errors,
    }
